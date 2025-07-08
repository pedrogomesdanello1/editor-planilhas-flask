# app.py (versão com ajuste de largura de colunas)

import flask
from flask import Flask, render_template, request, send_file, flash, redirect, url_for
import pandas as pd
import io
import os
import zipfile
import openpyxl
from openpyxl.utils import get_column_letter # Ferramenta para obter a letra da coluna (A, B, C...)
# No início do arquivo app.py
from openpyxl.styles import PatternFill, Font, Alignment

# Cria a aplicação web com Flask
app = Flask(__name__)
app.secret_key = 'chave-com-ajuste-de-largura'

@app.route('/')
def index():
    """Renderiza a página inicial (o arquivo index.html)."""
    return render_template('index.html')

@app.route('/processar', methods=['POST'])
def processar_planilha():
    """Função principal que processa a planilha, agora com formatação de largura."""
    if 'planilha' not in request.files:
        flash('Nenhum arquivo selecionado!')
        return redirect(url_for('index'))

    file = request.files['planilha']
    numeros_reais_str = request.form.get('numeros_reais')

    if file.filename == '' or not numeros_reais_str:
        flash('Por favor, envie a planilha e insira os números reais.')
        return redirect(url_for('index'))
    
    try:
        numeros_reais = int(numeros_reais_str)
        filename, file_extension = os.path.splitext(file.filename)
        
        df = None
        planilha_original_nome = file.filename

        if file_extension.lower() == '.zip':
             with zipfile.ZipFile(file, 'r') as zf:
                nomes_de_planilhas = [n for n in zf.namelist() if n.lower().endswith(('.xlsx', '.xls')) and not n.startswith('__MACOSX')]
                if not nomes_de_planilhas:
                    flash('O arquivo .zip enviado não contém nenhuma planilha Excel (.xlsx).')
                    return redirect(url_for('index'))
                planilha_original_nome = nomes_de_planilhas[0]
                with zf.open(planilha_original_nome) as planilha_interna:
                    df = pd.read_excel(planilha_interna)
        elif file_extension.lower() in ['.xlsx', '.xls']:
            df = pd.read_excel(file)
        else:
            flash(f"Formato de arquivo não suportado: {file_extension}. Por favor, use .xlsx ou .zip.")
            return redirect(url_for('index'))

        # Armazena os nomes originais das colunas antes de padronizar para minúsculo
        colunas_originais = df.columns.tolist()

        df.columns = [col.strip().lower() for col in df.columns]
        colunas_necessarias = ['veículos', 'data', 'url veiculada', 'impressões totais', 'categoria']
        if not all(coluna in df.columns for coluna in colunas_necessarias):
            flash(f"ERRO: As colunas essenciais {colunas_necessarias} não foram encontradas.")
            flash(f"Colunas encontradas (após limpeza): {df.columns.tolist()}")
            return redirect(url_for('index'))
        
        df['impressões totais'] = pd.to_numeric(df['impressões totais'], errors='coerce').fillna(0)
        impressoes_originais = df['impressões totais'].sum()
        diferenca_a_remover = int(impressoes_originais - numeros_reais)

        if diferenca_a_remover < 0:
            flash(f"O número real ({numeros_reais}) é maior que o total da planilha ({impressoes_originais}).")
            return redirect(url_for('index'))
        
        if diferenca_a_remover > 0:
            df.sort_values(by='impressões totais', ascending=False, inplace=True)
            df['remocao_potencial'] = df['impressões totais'] - 1
            df.loc[df['remocao_potencial'] < 0, 'remocao_potencial'] = 0
            df['remocao_acumulada'] = df['remocao_potencial'].cumsum()
            ponto_de_corte_series = df[df['remocao_acumulada'] >= diferenca_a_remover]
            if not ponto_de_corte_series.empty:
                idx_corte = ponto_de_corte_series.index[0]
                removido_antes = df.loc[:idx_corte, 'remocao_acumulada'].iloc[-2] if idx_corte != df.index[0] else 0
                restante_a_remover_na_linha_corte = diferenca_a_remover - removido_antes
                indices_antes_corte = df.loc[:idx_corte].index[:-1]
                df.loc[indices_antes_corte, 'impressões totais'] = 1
                df.loc[idx_corte, 'impressões totais'] -= restante_a_remover_na_linha_corte
            df.drop(columns=['remocao_potencial', 'remocao_acumulada'], inplace=True)
            df = df[df['impressões totais'] > 0]
        
        df['impressões totais'] = df['impressões totais'].astype(int)

        # Renomeia as colunas de volta para o formato original para a saída
        df.columns = colunas_originais

        output = io.BytesIO()
        nome_base, _ = os.path.splitext(planilha_original_nome)
        novo_nome_arquivo = f"{nome_base}.xlsx"

        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Sheet1')
            
            # --- INÍCIO DO BLOCO DE FORMATAÇÃO ---
            workbook = writer.book
            worksheet = writer.sheets['Sheet1']

            # 1. Dicionário de larguras (com chaves em minúsculo para correspondência)
            # Usei os nomes das colunas originais que o usuário passou, mas a lógica usa minúsculas
            col_widths = {'Veículos': 25.43, 'Data': 10.00, 'URL Veiculada': 100.00, 'Impressões Totais': 20.86, 'Categoria': 24.29}

            # 2. Mapeia o nome da coluna para sua letra (A, B, C...) e aplica a largura
            # df.columns aqui já contém os nomes originais (ex: 'Impressões Totais')
            for i, column_title in enumerate(df.columns):
                if column_title in col_widths:
                    column_letter = get_column_letter(i + 1)
                    worksheet.column_dimensions[column_letter].width = col_widths[column_title]

                    # 3. Define e aplica os estilos para o cabeçalho
            header_fill = PatternFill(start_color="47a432", end_color="47a432", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF") # Fonte branca em negrito para contraste
            header_alignment = Alignment(horizontal="center", vertical="center")

            # 4. Aplica os estilos a cada célula da primeira linha (o cabeçalho)
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
            # --- FIM DO BLOCO DE FORMATAÇÃO ---

        output.seek(0)

        return send_file(
            output, as_attachment=True, download_name=novo_nome_arquivo,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    except Exception as e:
        flash(f"Ocorreu um erro crítico: {e}")
        app.logger.error(f"Erro detalhado: {e}", exc_info=True)
        return redirect(url_for('index'))

if __name__ == '__main__':
    app.run(debug=True)
